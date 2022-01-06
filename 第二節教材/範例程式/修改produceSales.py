import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

for rowNum in range(2, sheet.max_row+1): # 跳過標頭行
    produceName = sheet.cell(row=rowNum, column=1).value
    
    if produceName == '大蒜':
        sheet.cell(row=rowNum, column=2).value = 3.07

wb.save('updatedProduceSales.xlsx')


