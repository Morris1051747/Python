import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.active 
#sheet = 工作表

print(sheet.max_row) 
print(sheet.max_column)
# row是橫的 column是直的

print(sheet['A1'].value)

c = sheet['B1']

print(c.value)


for i in range(1, 8):
	print(sheet.cell(row=i, column=2).value)
